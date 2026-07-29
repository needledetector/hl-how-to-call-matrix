#!/usr/bin/env python3
"""
呼称表 → 正規化JSON (v2)

  python convert.py 呼称表.xlsx -o data.json [--axis 軸マッピング.tsv] [--ids ids.json]

実データから確認した構造:
  * セルは4状態を持つ。呼称列 / 未確認(セル全体が「※」) /
    不成立(セル全体が括弧だけ。在籍期間が重ならない等) / 未着手(空欄)。
  * 「、」はセル内デリミタだが、括弧・鉤括弧・引用符の中にも現れる。深さを見て切る。
  * 記号は「ラベル(注記)記号」の順に付く:  俺(♂)*
  * 「←」はセルを二分する境界。左が現役の呼称、右が使用終了。個々の呼称の
    対応関係ではない。セル内の並び順は頻度順（先頭ほどよく使われる）。
  * 使用終了はもう一つ「(旧)」タグでも表される。←を伴わないセルが19件ある。
    左側に(旧)が来ることはないので、両者は同じことの別表記。
  * 半角カナはラベル側では早口・裏声などのニュアンスを担うので保存する。
    注記側だけ全角に寄せ、検索キーは別に正規化して持つ。
"""
import argparse, json, re, sys, unicodedata
from pathlib import Path

MARK_SUFFIX = "◎*+☆"
FLAG_OF = {"◎": "main", "*": "rare", "+": "third", "☆": "egosa"}
TIMECODE = re.compile(r"\d{1,2}:\d{2}(?::\d{2})?")
CELL_UNSURE = "※"
# 括弧だけのセルの中身のうち、「呼称が存在しない」を意味するもの
NA_REASONS = {"在籍時未デビュー", "デビュー時離籍済", "登場時離籍済",
              "活動中未登場", "活動中未デビュー", "在籍時未登場"}
OMITTED = {"省略"}


def to_fullwidth(s: str) -> str:
    return unicodedata.normalize("NFKC", s)


def search_key(s: str) -> str:
    return to_fullwidth(s).casefold().replace(" ", "").replace("　", "")


def split_outside(s: str, sep: str = "、") -> list[str]:
    """括弧・鉤括弧・引用符の外側にある区切り文字だけで分割する。"""
    out, buf, depth, quoted = [], [], 0, False
    for ch in s:
        if ch == '"':
            quoted = not quoted
        elif ch in "（(「『":
            depth += 1
        elif ch in "）)」』":
            depth = max(0, depth - 1)
        if ch == sep and depth == 0 and not quoted:
            out.append("".join(buf)); buf = []
        else:
            buf.append(ch)
    out.append("".join(buf))
    return [t.strip() for t in out if t.strip()]


WHOLE_PAREN = re.compile(r"^[（(](.*)[）)]$", re.S)


LEAD_PAREN = re.compile(r"^[（(]([^）)]*)[）)]\s*")


def cell_state(raw: str) -> tuple[str | None, str | None, str]:
    """セル先頭の括弧を関係の注記として剥がし、(状態, 理由, 残り) を返す。
    離籍後でも三人称でなら言及できるので、注記と呼称は共存しうる。
      例: (デビュー時離籍済) 会長先輩+
    """
    s = raw.strip()
    if s == CELL_UNSURE:
        return "unsure", None, ""
    m = LEAD_PAREN.match(s)
    if m:
        inner = m.group(1).strip().strip("「」『』")
        rest = s[m.end():].strip()
        if inner in OMITTED:
            return "omitted", inner, rest
        if inner in NA_REASONS or "離籍" in inner or "未デビュー" in inner or "卒業" in inner:
            return "na", inner, rest
    return None, None, s


def parse_appellation(part: str, flags: dict) -> dict | None:
    t = part.strip()
    notes: list[str] = []
    for _ in range(4):
        before = t
        m = re.search(rf"([{re.escape(MARK_SUFFIX)}]+)\s*$", t)
        if m:
            for ch in m.group(1):
                flags[FLAG_OF[ch]] = True
            t = t[: m.start()].rstrip()
        m = re.search(r"[（(]([^）)]*)[）)]\s*$", t)
        if m:
            notes.insert(0, m.group(1)); t = t[: m.start()].rstrip()
        if t == before:
            break
    if not t and not notes:
        return None

    tags, times = [], []
    for n in notes:
        for piece in split_outside(to_fullwidth(n)):
            for tag in re.split(r"[、/／]", piece):
                tag = tag.strip()
                if not tag:
                    continue
                if TIMECODE.search(tag):
                    times.append(tag)
                tags.append(tag)
    return {"label": t, "key": search_key(t), "tags": tags, "times": times}


def parse_token(raw: str) -> dict | None:
    flags = {k: False for k in ("main", "rare", "third", "egosa")}
    a = parse_appellation(raw, flags)
    if a is None:
        return None
    a["flags"] = [k for k, v in flags.items() if v]
    return a


def parse_cell(raw: str) -> list[dict]:
    """セルを ← で区切り、左を現役・右を使用終了として並べる。順序は保つ。"""
    out = []
    for gi, seg in enumerate(re.split(r"←|<-", raw)):
        for tok in split_outside(seg):
            a = parse_token(tok)
            if not a:
                continue
            a["retired"] = gi > 0 or "旧" in a["tags"]
            if a["retired"]:
                a["flags"] = a["flags"] + ["retired"]
            out.append(a)
    return out


def read_workbook(path: Path):
    import openpyxl
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["呼称表"]
    rows = [r for r in ws.iter_rows(values_only=True) if any(c is not None for c in r)]
    cols = [c for c in rows[0][1:] if c is not None]
    names = [r[0] for r in rows[1:]]
    matrix = [[str(r[1 + j]).replace("\n", " ") if r[1 + j] else ""
               for j in range(len(cols))] for r in rows[1:]]
    # 補助データは位置ではなく見出しで引く（列の増減で壊れないように）
    aux = {}
    if "補助データ" in wb.sheetnames:
        arows = [r for r in wb["補助データ"].iter_rows(values_only=True)
                 if r and any(c is not None for c in r)]
        if arows:
            head = [str(c).strip() if c is not None else "" for c in arows[0]]

            def col(*cands):
                for n in cands:
                    if n in head:
                        return head.index(n)
                return -1

            ci = {"name": col("人物", "名前", "キャラ", "キャラクター"),
                  "id": col("id", "ID", "Id"),
                  "proj": col("グループ", "プロジェクト"),
                  "gen": col("期生"), "gen2": col("期生兼"),
                  "emo": col("絵文字"),
                  "abbr": col("略称", "略", "短縮名", "短縮")}

            def get(r, i):
                if i < 0 or i >= len(r) or r[i] is None:
                    return None
                v = str(r[i]).strip()
                return v or None

            for r in arows[1:]:
                nm = get(r, ci["name"] if ci["name"] >= 0 else 0)
                if not nm:
                    continue
                aux[nm] = {"project": get(r, ci["proj"]),
                           "gens": [g for g in (get(r, ci["gen"]), get(r, ci["gen2"])) if g],
                           "emoji": get(r, ci["emo"]),
                           "abbr": get(r, ci["abbr"]),
                           "id": get(r, ci["id"])}
    return names, cols, matrix, aux, rows[0][0]


def load_axes(path: Path | None) -> dict:
    if not path or not path.exists():
        return {}
    out = {}
    for line in path.read_text(encoding="utf-8").splitlines():
        if not line.strip() or line.startswith("#"):
            continue
        f = line.split("\t")
        if len(f) >= 2:
            out[to_fullwidth(f[0].strip())] = {
                "axis": f[1].strip(),
                "label": f[2].strip() if len(f) > 2 and f[2].strip() else f[0].strip()}
    return out


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("src", type=Path)
    ap.add_argument("-o", "--out", type=Path, default=Path("data.json"))
    ap.add_argument("--axis", type=Path, default=Path("軸マッピング.tsv"))
    ap.add_argument("--ids", type=Path, default=Path("ids.json"))
    args = ap.parse_args()

    names, cols, matrix, aux, legend = read_workbook(args.src)
    axes = load_axes(args.axis)

    id_map = json.loads(args.ids.read_text(encoding="utf-8")) if args.ids.exists() else {}
    nxt = max([int(v[1:]) for v in id_map.values() if v[1:].isdigit()] or [0]) + 1
    for n in names:
        sheet_id = (aux.get(n) or {}).get("id")
        if sheet_id:
            id_map[n] = sheet_id          # シートに id 列があればそちらを優先
        elif n not in id_map:
            id_map[n] = f"c{nxt:03d}"; nxt += 1
    args.ids.write_text(json.dumps(id_map, ensure_ascii=False, indent=1), encoding="utf-8")

    if cols != names:
        diff = [(a, b) for a, b in zip(names, cols) if a != b]
        print(f"警告: 行と列の見出しが {len(diff)} 件ずれています → {diff[:3]}", file=sys.stderr)

    chars = [{"id": id_map[n], "name": n, "key": search_key(n),
              "emoji": aux.get(n, {}).get("emoji"),
              "abbr": aux.get(n, {}).get("abbr"),
              "project": aux.get(n, {}).get("project"),
              "gens": aux.get(n, {}).get("gens", []),
              "known": n in aux} for n in names]

    cells, unmapped, stats = [], {}, {"has": 0, "unsure": 0, "na": 0, "omitted": 0, "empty": 0}
    both = 0
    n_app = 0
    for i, row in enumerate(matrix):
        for j, raw in enumerate(row):
            f, t = id_map[names[i]], id_map[names[j]]
            if not raw.strip():
                stats["empty"] += 1
                continue
            state, reason, rest = cell_state(raw)
            stats[state or "has"] += 1
            cell = {"f": f, "t": t}
            if state:
                cell["s"] = state
                if reason:
                    cell["r"] = reason
            apps = []
            for tok in parse_cell(rest) if rest else []:
                ax = {}
                for tag in tok["tags"]:
                    hit = axes.get(tag)
                    if hit:
                        ax.setdefault(hit["axis"], []).append(hit["label"])
                    else:
                        unmapped[tag] = unmapped.get(tag, 0) + 1
                a = {"l": tok["label"]}
                if tok["key"] != tok["label"]:
                    a["k"] = tok["key"]
                if tok["flags"]:
                    a["g"] = tok["flags"]
                if tok["tags"]:
                    a["n"] = tok["tags"]
                if ax:
                    a["x"] = ax
                if tok["times"]:
                    a["src"] = tok["times"]
                apps.append(a)
                n_app += 1
            if apps:
                cell["a"] = apps
            if "s" in cell or "a" in cell:
                cells.append(cell)

    out = {"version": 2, "legend": legend, "chars": chars, "cells": cells,
           "axes": {k: v for k, v in axes.items()}}
    args.out.write_text(json.dumps(out, ensure_ascii=False, separators=(",", ":")),
                        encoding="utf-8")

    print(f"キャラ {len(chars)}  所属不明 {sum(1 for c in chars if not c['known'])}")
    both = sum(1 for c in cells if "s" in c and "a" in c)
    print("セル状態: " + "  ".join(f"{k}={v}" for k, v in stats.items()) +
          f"  （うち注記と呼称の共存 {both}）")
    print(f"呼称 {n_app}")
    fc = {k: sum(1 for c in cells for a in c.get("a", []) if k in a.get("g", []))
          for k in ("main", "rare", "third", "egosa", "retired")}
    print("  フラグ: " + "  ".join(f"{k}={v}" for k, v in fc.items()))
    mapped = sum(1 for c in cells for a in c.get("a", []) if "x" in a)
    print(f"軸マッピング済み {mapped} 件 / 未マッピング {len(unmapped)} 種 {sum(unmapped.values())} 件")
    if unmapped:
        top = sorted(unmapped.items(), key=lambda x: -x[1])[:12]
        print("  未マッピング上位: " + "  ".join(f"{k}({v})" for k, v in top))
    if fc["egosa"] == 0:
        print("  注意: ☆(エゴサワード)が0件です。シートが未反映かもしれません。", file=sys.stderr)
    print(f"出力 {args.out}  {args.out.stat().st_size / 1024:.0f}KB")


if __name__ == "__main__":
    main()
